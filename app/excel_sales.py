from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
import gc
import os
from pathlib import Path
import shutil
import tempfile
from typing import Any
from zipfile import is_zipfile

from app.config import SALES_EXCEL_DATE_FORMAT, SALES_WORKBOOK_PATH, SALES_WORKSHEET_NAME


@dataclass(frozen=True)
class ExcelSyncResult:
    saved: bool
    row: int | None = None
    message: str = ""


class SalesWorkbook:
    CALCULATED_HEADERS = {
        "profit",
        "total expenses",
        "total profit",
        "net profit",
    }
    HEADERS = (
        "Customer Name",
        "Items Sold",
        "Email/Order ID",
        "Buying Amount",
        "Selling Amount",
        "Profit",
        "Status",
        "Date",
    )
    ENTRY_DATA_COLUMNS = (1, 2, 3, 4, 5, 7)
    SCREEN_SERVICE_LIMITS = {
        "netflix screen": 5,
        "netflix screens": 5,
        "hbo max screen": 9,
        "hbo max screens": 9,
    }

    def __init__(self, path: Path | str = SALES_WORKBOOK_PATH, worksheet_name: str = SALES_WORKSHEET_NAME) -> None:
        self.target = os.path.expandvars(str(path)).strip()
        self.path = Path(self.target).expanduser()
        self.worksheet_name = worksheet_name

    @property
    def display_path(self) -> str:
        return self.target

    @property
    def is_cloud_url(self) -> bool:
        return self.target.lower().startswith(("http://", "https://"))

    def sync_entry(self, entry: dict[str, Any]) -> ExcelSyncResult:
        if self._prefer_excel_com():
            com_result = self._sync_entry_with_excel_com(entry)
            if com_result is not None:
                return com_result

        workbook = None
        row: int | None = None
        try:
            row = self._existing_row(entry)
            matched_retry_row = False
            workbook, sheet = self._open_workbook()
            self._ensure_headers(sheet)
            screen_result = self._sync_screen_entry(sheet, entry)
            if screen_result is not None:
                if screen_result.saved:
                    self._request_formula_recalculation(workbook)
                    self._save_workbook(workbook)
                return screen_result
            if row is not None and (row > sheet.max_row or not self._row_identity_matches_entry(sheet, row, entry)):
                row = None
            if row is None:
                row = self._matching_retry_row(sheet, entry)
                matched_retry_row = row is not None
            if matched_retry_row:
                return ExcelSyncResult(True, row=row, message=f"Existing Excel row {row}")
            if row is None:
                row = self._next_row(sheet)
            self._write_row(sheet, row, entry)
            self._request_formula_recalculation(workbook)
            self._save_workbook(workbook)
        except Exception as exc:
            recovered_row = self._recover_saved_row(entry, row)
            if recovered_row is not None:
                return ExcelSyncResult(True, row=recovered_row, message=f"Excel row {recovered_row}")
            com_result = self._sync_entry_with_excel_com(entry)
            if com_result is not None:
                return com_result
            return ExcelSyncResult(False, message=self._sync_error_message(exc))
        finally:
            self._close_workbook(workbook)
        return ExcelSyncResult(True, row=row, message=f"Excel row {row}")

    def delete_row(self, row: int) -> ExcelSyncResult:
        if row <= 1:
            return ExcelSyncResult(False, message="Excel data row is invalid.")
        workbook = None
        try:
            workbook, sheet = self._open_workbook()
            if row > sheet.max_row:
                return ExcelSyncResult(False, row=row, message=f"Excel row {row} does not exist.")
            sheet.delete_rows(row, 1)
            self._request_formula_recalculation(workbook)
            self._save_workbook(workbook)
        except Exception as exc:
            return ExcelSyncResult(False, row=row, message=self._sync_error_message(exc))
        finally:
            self._close_workbook(workbook)
        return ExcelSyncResult(True, row=row, message=f"Excel row {row} removed")

    def _existing_row(self, entry: dict[str, Any]) -> int | None:
        value = entry.get("excel_row")
        if value in (None, ""):
            return None
        try:
            row = int(value)
        except (TypeError, ValueError):
            return None
        return row if row > 1 else None

    def _matching_retry_row(self, sheet, entry: dict[str, Any]) -> int | None:
        if not entry.get("excel_sync_error"):
            return None
        calculated_columns = self._calculated_columns(sheet)
        for row in range(max(sheet.max_row, 1), 1, -1):
            if not self._row_has_entered_data(sheet, row, calculated_columns):
                continue
            if self._row_matches_entry(sheet, row, entry):
                return row
        return None

    def _recover_saved_row(self, entry: dict[str, Any], expected_row: int | None) -> int | None:
        if self.is_cloud_url or not self.path.exists():
            return None
        workbook = None
        try:
            from openpyxl import load_workbook

            workbook = load_workbook(self.path, keep_vba=self.path.suffix.lower() == ".xlsm")
            sheet = self._select_sheet(workbook)
            if expected_row is not None and expected_row <= sheet.max_row and self._row_matches_entry(sheet, expected_row, entry):
                return expected_row
            retry_entry = dict(entry)
            retry_entry["excel_sync_error"] = retry_entry.get("excel_sync_error") or "Checking for saved Excel row."
            return self._matching_retry_row(sheet, retry_entry)
        except Exception:
            return None
        finally:
            self._close_workbook(workbook)

    def _row_matches_entry(self, sheet, row: int, entry: dict[str, Any]) -> bool:
        comparisons = (
            (1, entry.get("customer", ""), "text"),
            (2, entry.get("item", ""), "text"),
            (3, entry.get("order_id", ""), "text"),
            (4, entry.get("buying_amount", "0"), "number"),
            (5, entry.get("selling_amount", ""), "number"),
            (7, entry.get("status", ""), "text"),
        )
        for column, expected, kind in comparisons:
            actual = sheet.cell(row=row, column=column).value
            if kind == "number":
                if self._number_key(actual) != self._number_key(expected):
                    return False
            elif self._text_key(actual) != self._text_key(expected):
                return False
        return True

    def _text_key(self, value: Any) -> str:
        return " ".join(str(value or "").split()).casefold()

    def _number_key(self, value: Any) -> str:
        converted = self._number_or_text(value)
        if isinstance(converted, (int, float)):
            return f"{float(converted):.6f}".rstrip("0").rstrip(".")
        return self._text_key(converted)

    def _screen_limit(self, entry: dict[str, Any]) -> int | None:
        item_key = self._text_key(entry.get("item", ""))
        return self.SCREEN_SERVICE_LIMITS.get(item_key)

    def _customer_names(self, value: Any) -> list[str]:
        text = str(value or "").strip()
        if not text:
            return []
        return [part.strip() for part in text.split(",") if part.strip()]

    def _customer_exists(self, customers: list[str], customer: str) -> bool:
        customer_key = self._text_key(customer)
        return any(self._text_key(existing) == customer_key for existing in customers)

    def _customer_index(self, customers: list[str], customer: str) -> int | None:
        customer_key = self._text_key(customer)
        for index, existing in enumerate(customers):
            if self._text_key(existing) == customer_key:
                return index
        return None

    def _replace_customer(self, customers: list[str], old_customer: str, new_customer: str) -> list[str] | None:
        old_index = self._customer_index(customers, old_customer)
        if old_index is None:
            return None
        new_index = self._customer_index(customers, new_customer)
        if new_index is not None and new_index != old_index:
            return [customer for index, customer in enumerate(customers) if index != old_index]
        updated = list(customers)
        updated[old_index] = new_customer
        return updated

    def _screen_full_message(self, entry: dict[str, Any], limit: int) -> str:
        service = str(entry.get("item") or "This screen service").strip()
        return f"{service} account is full ({limit} customers). Please use a new account email."

    def _add_numbers(self, current: Any, added: Any) -> int | float | Any:
        current_number = self._number_or_text(current)
        added_number = self._number_or_text(added)
        if self._is_blank(current) and isinstance(added_number, (int, float)):
            return added_number
        if not isinstance(current_number, (int, float)) or not isinstance(added_number, (int, float)):
            return current
        total = float(current_number) + float(added_number)
        if total.is_integer():
            return int(total)
        return total

    def _screen_row_matches_values(self, row_item: Any, row_order_id: Any, item: Any, order_id: Any) -> bool:
        return (
            self._text_key(row_item) == self._text_key(item)
            and self._text_key(row_order_id) == self._text_key(order_id)
        )

    def _find_screen_row_for_values(self, sheet, item: Any, order_id: Any) -> int | None:
        if not self._text_key(order_id):
            return None
        for row in range(max(sheet.max_row, 1), 1, -1):
            if self._screen_row_matches_values(
                sheet.cell(row=row, column=2).value,
                sheet.cell(row=row, column=3).value,
                item,
                order_id,
            ):
                return row
        return None

    def _find_screen_row(self, sheet, entry: dict[str, Any]) -> int | None:
        return self._find_screen_row_for_values(sheet, entry.get("item", ""), entry.get("order_id", ""))

    def _sync_screen_entry(self, sheet, entry: dict[str, Any]) -> ExcelSyncResult | None:
        limit = self._screen_limit(entry)
        if limit is None:
            return None
        previous_customer = str(entry.get("previous_customer") or "").strip()
        previous_item = entry.get("previous_item") or entry.get("item", "")
        previous_order_id = entry.get("previous_order_id") or entry.get("order_id", "")
        row = self._find_screen_row_for_values(sheet, previous_item, previous_order_id)
        if row is None:
            row = self._find_screen_row(sheet, entry)
        if row is None:
            return None

        customer = str(entry.get("customer") or "").strip()
        customers = self._customer_names(sheet.cell(row=row, column=1).value)
        if previous_customer:
            replaced = self._replace_customer(customers, previous_customer, customer)
            if replaced is not None:
                sheet.cell(row=row, column=1).value = ", ".join(replaced)
                sheet.cell(row=row, column=6).value = f"=E{row}-D{row}"
                return ExcelSyncResult(True, row=row, message=f"Excel row {row}")
        if self._customer_exists(customers, customer):
            return ExcelSyncResult(True, row=row, message=f"Existing Excel row {row}")
        if len(customers) >= limit:
            return ExcelSyncResult(False, row=row, message=self._screen_full_message(entry, limit))

        customers.append(customer)
        sheet.cell(row=row, column=1).value = ", ".join(customers)
        sheet.cell(row=row, column=5).value = self._add_numbers(
            sheet.cell(row=row, column=5).value,
            entry.get("selling_amount", ""),
        )
        sheet.cell(row=row, column=6).value = f"=E{row}-D{row}"
        return ExcelSyncResult(True, row=row, message=f"Excel row {row}")

    def _prefer_excel_com(self) -> bool:
        return os.name == "nt" and (
            self.is_cloud_url or any("onedrive" in part.casefold() for part in self.path.parts)
        )

    def _sync_entry_with_excel_com(self, entry: dict[str, Any]) -> ExcelSyncResult | None:
        if os.name != "nt" or (not self.is_cloud_url and not self.path.exists()):
            return None

        try:
            import pythoncom
            import win32com.client
        except ModuleNotFoundError:
            return None

        pythoncom.CoInitialize()
        excel = None
        workbook = None
        try:
            excel = win32com.client.DispatchEx("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            excel.EnableEvents = False
            excel.ScreenUpdating = False

            workbook = excel.Workbooks.Open(
                self.target,
                UpdateLinks=0,
                ReadOnly=False,
                Notify=False,
                AddToMru=False,
                IgnoreReadOnlyRecommended=True,
            )
            if bool(getattr(workbook, "ReadOnly", False)):
                raise PermissionError("Sales workbook opened read-only.")

            sheet = self._select_com_sheet(workbook)
            self._ensure_com_headers(sheet)
            screen_result = self._sync_screen_entry_com(sheet, entry)
            if screen_result is not None:
                if screen_result.saved:
                    workbook.Save()
                return screen_result
            row = self._existing_row(entry)
            matched_retry_row = False
            max_row, _ = self._com_used_bounds(sheet)
            if row is not None and (row > max_row or not self._com_row_identity_matches_entry(sheet, row, entry)):
                row = None
            if row is None:
                row = self._matching_retry_row_com(sheet, entry)
                matched_retry_row = row is not None
            if matched_retry_row:
                return ExcelSyncResult(True, row=row, message=f"Existing Excel row {row}")
            if row is None:
                row = self._next_row_com(sheet)
            self._write_com_row(sheet, row, entry)
            workbook.Save()
            return ExcelSyncResult(True, row=row, message=f"Excel row {row}")
        except Exception as exc:
            return ExcelSyncResult(False, message=self._sync_error_message(exc))
        finally:
            if workbook is not None:
                try:
                    workbook.Close(SaveChanges=False)
                except Exception:
                    pass
            if excel is not None:
                try:
                    excel.Quit()
                except Exception:
                    pass
            pythoncom.CoUninitialize()

    def _select_com_sheet(self, workbook):
        if not self.worksheet_name:
            return workbook.ActiveSheet
        for index in range(1, workbook.Worksheets.Count + 1):
            sheet = workbook.Worksheets(index)
            if str(sheet.Name) == self.worksheet_name:
                return sheet
        sheet = workbook.Worksheets.Add(After=workbook.Worksheets(workbook.Worksheets.Count))
        sheet.Name = self.worksheet_name
        return sheet

    def _ensure_com_headers(self, sheet) -> None:
        has_headers = any(not self._is_blank(sheet.Cells(1, column).Value) for column in range(1, 9))
        if has_headers:
            return
        widths = (22, 26, 26, 16, 16, 14, 14, 14)
        for column, header in enumerate(self.HEADERS, start=1):
            sheet.Cells(1, column).Value = header
            sheet.Columns(column).ColumnWidth = widths[column - 1]

    def _matching_retry_row_com(self, sheet, entry: dict[str, Any]) -> int | None:
        if not entry.get("excel_sync_error"):
            return None
        max_row, max_column = self._com_used_bounds(sheet)
        calculated_columns = self._com_calculated_columns(sheet, max_column)
        for row in range(max(max_row, 1), 1, -1):
            if not self._com_row_has_entered_data(sheet, row, max_column, calculated_columns):
                continue
            if self._com_row_matches_entry(sheet, row, entry):
                return row
        return None

    def _next_row_com(self, sheet) -> int:
        max_row, max_column = self._com_used_bounds(sheet)
        calculated_columns = self._com_calculated_columns(sheet, max_column)
        for row in range(max(max_row, 1), 1, -1):
            if self._com_row_has_entered_data(sheet, row, max_column, calculated_columns):
                return row + 1
        return 2

    def _com_used_bounds(self, sheet) -> tuple[int, int]:
        used_range = sheet.UsedRange
        max_row = int(used_range.Row + used_range.Rows.Count - 1)
        max_column = int(used_range.Column + used_range.Columns.Count - 1)
        return max(max_row, 1), max(max_column, len(self.HEADERS))

    def _com_calculated_columns(self, sheet, max_column: int) -> set[int]:
        columns = set()
        for column in range(1, max(max_column, len(self.HEADERS)) + 1):
            header = self._cell_text(sheet.Cells(1, column).Value).lower()
            if header in self.CALCULATED_HEADERS:
                columns.add(column)
        return columns

    def _com_row_has_entered_data(self, sheet, row: int, max_column: int, calculated_columns: set[int]) -> bool:
        for column in self.ENTRY_DATA_COLUMNS:
            cell = sheet.Cells(row, column)
            if self._is_blank(cell.Value):
                continue
            return True
        return False

    def _com_row_identity_matches_entry(self, sheet, row: int, entry: dict[str, Any]) -> bool:
        return all(
            self._text_key(sheet.Cells(row, column).Value) == self._text_key(entry.get(key, ""))
            for column, key in ((1, "customer"), (2, "item"), (3, "order_id"))
        )

    def _find_screen_row_com_for_values(self, sheet, item: Any, order_id: Any) -> int | None:
        if not self._text_key(order_id):
            return None
        max_row, _ = self._com_used_bounds(sheet)
        for row in range(max(max_row, 1), 1, -1):
            if self._screen_row_matches_values(
                sheet.Cells(row, 2).Value,
                sheet.Cells(row, 3).Value,
                item,
                order_id,
            ):
                return row
        return None

    def _find_screen_row_com(self, sheet, entry: dict[str, Any]) -> int | None:
        return self._find_screen_row_com_for_values(sheet, entry.get("item", ""), entry.get("order_id", ""))

    def _sync_screen_entry_com(self, sheet, entry: dict[str, Any]) -> ExcelSyncResult | None:
        limit = self._screen_limit(entry)
        if limit is None:
            return None
        previous_customer = str(entry.get("previous_customer") or "").strip()
        previous_item = entry.get("previous_item") or entry.get("item", "")
        previous_order_id = entry.get("previous_order_id") or entry.get("order_id", "")
        row = self._find_screen_row_com_for_values(sheet, previous_item, previous_order_id)
        if row is None:
            row = self._find_screen_row_com(sheet, entry)
        if row is None:
            return None

        customer = str(entry.get("customer") or "").strip()
        customers = self._customer_names(sheet.Cells(row, 1).Value)
        if previous_customer:
            replaced = self._replace_customer(customers, previous_customer, customer)
            if replaced is not None:
                sheet.Cells(row, 1).Value = ", ".join(replaced)
                sheet.Cells(row, 6).Formula = f"=E{row}-D{row}"
                return ExcelSyncResult(True, row=row, message=f"Excel row {row}")
        if self._customer_exists(customers, customer):
            return ExcelSyncResult(True, row=row, message=f"Existing Excel row {row}")
        if len(customers) >= limit:
            return ExcelSyncResult(False, row=row, message=self._screen_full_message(entry, limit))

        customers.append(customer)
        sheet.Cells(row, 1).Value = ", ".join(customers)
        sheet.Cells(row, 5).Value = self._add_numbers(sheet.Cells(row, 5).Value, entry.get("selling_amount", ""))
        sheet.Cells(row, 6).Formula = f"=E{row}-D{row}"
        return ExcelSyncResult(True, row=row, message=f"Excel row {row}")

    def _com_row_matches_entry(self, sheet, row: int, entry: dict[str, Any]) -> bool:
        comparisons = (
            (1, entry.get("customer", ""), "text"),
            (2, entry.get("item", ""), "text"),
            (3, entry.get("order_id", ""), "text"),
            (4, entry.get("buying_amount", "0"), "number"),
            (5, entry.get("selling_amount", ""), "number"),
            (7, entry.get("status", ""), "text"),
        )
        for column, expected, kind in comparisons:
            actual = sheet.Cells(row, column).Value
            if kind == "number":
                if self._number_key(actual) != self._number_key(expected):
                    return False
            elif self._text_key(actual) != self._text_key(expected):
                return False
        return True

    def _write_com_row(self, sheet, row: int, entry: dict[str, Any]) -> None:
        sheet.Cells(row, 1).Value = entry.get("customer", "")
        sheet.Cells(row, 2).Value = entry.get("item", "")
        sheet.Cells(row, 3).Value = entry.get("order_id", "")
        sheet.Cells(row, 4).Value = self._number_or_text(entry.get("buying_amount", "0"))
        sheet.Cells(row, 5).Value = self._number_or_text(entry.get("selling_amount", ""))
        sheet.Cells(row, 6).Formula = f"=E{row}-D{row}"
        sheet.Cells(row, 7).Value = entry.get("status", "")
        date_value = self._date_value(entry)
        if isinstance(date_value, datetime):
            date_value = date_value.date()
        if isinstance(date_value, date):
            date_value = self._excel_date_serial(date_value, bool(getattr(sheet.Parent, "Date1904", False)))
        sheet.Cells(row, 8).Value = date_value
        sheet.Cells(row, 8).NumberFormat = SALES_EXCEL_DATE_FORMAT

    def _open_workbook(self):
        if self.is_cloud_url:
            raise ValueError("Cloud workbook URLs require Microsoft Excel sync on Windows.")
        if self.path.suffix.lower() not in {".xlsx", ".xlsm"}:
            raise ValueError("Sales workbook must be an .xlsx or .xlsm file.")

        try:
            from openpyxl import Workbook, load_workbook
        except ModuleNotFoundError as exc:
            raise RuntimeError("openpyxl is not installed. Run: pip install -r requirements.txt") from exc

        self.path.parent.mkdir(parents=True, exist_ok=True)
        if self.path.exists():
            workbook = load_workbook(self.path, keep_vba=self.path.suffix.lower() == ".xlsm")
            return workbook, self._select_sheet(workbook)

        workbook = Workbook()
        sheet = workbook.active
        if self.worksheet_name:
            sheet.title = self.worksheet_name
        return workbook, sheet

    def _select_sheet(self, workbook):
        if not self.worksheet_name:
            return workbook.active
        if self.worksheet_name in workbook.sheetnames:
            return workbook[self.worksheet_name]
        return workbook.create_sheet(self.worksheet_name)

    def _ensure_headers(self, sheet) -> None:
        has_headers = any(sheet.cell(row=1, column=column).value not in (None, "") for column in range(1, 9))
        if not has_headers:
            for column, header in enumerate(self.HEADERS, start=1):
                sheet.cell(row=1, column=column).value = header
            widths = (22, 26, 26, 16, 16, 14, 14, 14)
            for column, width in enumerate(widths, start=1):
                sheet.column_dimensions[chr(64 + column)].width = width

    def _next_row(self, sheet) -> int:
        calculated_columns = self._calculated_columns(sheet)
        for row in range(max(sheet.max_row, 1), 1, -1):
            if self._row_has_entered_data(sheet, row, calculated_columns):
                return row + 1
        return 2

    def _calculated_columns(self, sheet) -> set[int]:
        columns = set()
        for column in range(1, max(sheet.max_column, len(self.HEADERS)) + 1):
            header = self._cell_text(sheet.cell(row=1, column=column).value).lower()
            if header in self.CALCULATED_HEADERS:
                columns.add(column)
        return columns

    def _row_has_entered_data(self, sheet, row: int, calculated_columns: set[int]) -> bool:
        for column in self.ENTRY_DATA_COLUMNS:
            value = sheet.cell(row=row, column=column).value
            if self._is_blank(value):
                continue
            return True
        return False

    def _row_identity_matches_entry(self, sheet, row: int, entry: dict[str, Any]) -> bool:
        return all(
            self._text_key(sheet.cell(row=row, column=column).value) == self._text_key(entry.get(key, ""))
            for column, key in ((1, "customer"), (2, "item"), (3, "order_id"))
        )

    def _is_blank(self, value: Any) -> bool:
        return value is None or (isinstance(value, str) and not value.strip())

    def _is_formula(self, value: Any) -> bool:
        return isinstance(value, str) and value.lstrip().startswith("=")

    def _cell_text(self, value: Any) -> str:
        return str(value or "").strip()

    def _write_row(self, sheet, row: int, entry: dict[str, Any]) -> None:
        sheet.cell(row=row, column=1).value = entry.get("customer", "")
        sheet.cell(row=row, column=2).value = entry.get("item", "")
        sheet.cell(row=row, column=3).value = entry.get("order_id", "")
        sheet.cell(row=row, column=4).value = self._number_or_text(entry.get("buying_amount", "0"))
        sheet.cell(row=row, column=5).value = self._number_or_text(entry.get("selling_amount", ""))
        sheet.cell(row=row, column=6).value = f"=E{row}-D{row}"
        sheet.cell(row=row, column=7).value = entry.get("status", "")

        date_cell = sheet.cell(row=row, column=8)
        date_cell.value = self._date_value(entry)
        date_cell.number_format = SALES_EXCEL_DATE_FORMAT

    def _number_or_text(self, value: Any) -> int | float | str:
        text = str(value or "").strip()
        if not text:
            return ""
        normalized = text.replace(",", "")
        try:
            number = float(normalized)
        except ValueError:
            return text
        if number.is_integer():
            return int(number)
        return number

    def _date_value(self, entry: dict[str, Any]):
        raw_date = str(entry.get("date") or entry.get("entry_date") or "").strip()
        if raw_date:
            try:
                return datetime.strptime(raw_date, "%Y-%m-%d").date()
            except ValueError:
                return raw_date
        return datetime.now().date()

    def _excel_date_serial(self, value: date, date_1904: bool = False) -> int:
        epoch = date(1904, 1, 1) if date_1904 else date(1899, 12, 30)
        return (value - epoch).days

    def _request_formula_recalculation(self, workbook) -> None:
        try:
            workbook.calculation.fullCalcOnLoad = True
        except AttributeError:
            pass

    def _save_workbook(self, workbook) -> None:
        temporary_path = self._temporary_workbook_path()
        try:
            self._save_to_temporary_workbook(workbook, temporary_path)
            self._close_workbook(workbook)
            gc.collect()
            try:
                temporary_path.replace(self.path)
            except PermissionError:
                shutil.copyfile(temporary_path, self.path)
        finally:
            try:
                temporary_path.unlink(missing_ok=True)
            except OSError:
                pass

    def _temporary_workbook_path(self) -> Path:
        temporary = tempfile.NamedTemporaryFile(delete=False, suffix=self.path.suffix)
        temporary_path = Path(temporary.name)
        temporary.close()
        return temporary_path

    def _save_to_temporary_workbook(self, workbook, temporary_path: Path) -> None:
        try:
            workbook.save(temporary_path)
        except OSError as exc:
            if getattr(exc, "errno", None) == 9 and self._valid_workbook_file(temporary_path):
                return
            raise

    def _valid_workbook_file(self, path: Path) -> bool:
        try:
            return path.exists() and path.stat().st_size > 0 and is_zipfile(path)
        except OSError:
            return False

    def _close_workbook(self, workbook) -> None:
        if workbook is None:
            return
        close = getattr(workbook, "close", None)
        if callable(close):
            try:
                close()
            except (OSError, ValueError):
                pass
        archive = getattr(workbook, "_archive", None)
        archive_close = getattr(archive, "close", None)
        if callable(archive_close):
            try:
                archive_close()
            except (OSError, ValueError):
                pass

    def _sync_error_message(self, exc: Exception) -> str:
        if isinstance(exc, PermissionError):
            return (
                "Excel workbook could not be written. Close the workbook in Excel, "
                "make sure it is available in OneDrive, then try syncing again."
            )
        if isinstance(exc, OSError) and getattr(exc, "errno", None) == 9:
            return (
                "OneDrive closed the workbook file handle while saving. Make sure the workbook "
                "is kept on this device and is not open in Excel, then try again."
            )
        return str(exc)
